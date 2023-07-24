import pandas as pd
import os
import time
from data_manager import file_select
from scipy import constants
import openpyxl
import matplotlib.pyplot as plt


def energy_total(hcl, bucket, y, i_start, i_stop):
    return sum(bucket[i_start:i_stop]*y[i_start:i_stop]*(10**-6)*(1/hcl[i_start:i_stop]))


# Planck's Constant (Js)
h = constants.h
# Speed of Light (m/s)
c = constants.c

name, df = file_select()
df[df < 0] = 0
cols = df.columns

# Wavelength (m)
lam = df[cols[0]]

t = []
for i in range(0, len(lam)):
    if i < len(lam)-1:
        t.append(lam[i+1]-lam[i])
    else:
        t.append(t[i-1])

bucket = pd.Series(t)

hcl = (h*c) / (lam * (10**-9))

incid_i_start = lam.where(350 < lam).first_valid_index()
incid_i_stop = lam.where(lam < 400).last_valid_index()

emitt_i_start = lam.where(400 < lam).first_valid_index()
emitt_i_stop = lam.where(lam < 1000).last_valid_index()

blank_sum = energy_total(hcl, bucket, df[cols[1]], incid_i_start, incid_i_stop)
pl_sum = energy_total(hcl, bucket, df[cols[2]], incid_i_start, incid_i_stop)
pl_int_sum = energy_total(hcl, bucket, df[cols[3]], emitt_i_start, emitt_i_stop)

total_photons_absorbed = blank_sum - pl_sum
perc_absorbed = total_photons_absorbed / blank_sum
plqy = pl_int_sum / total_photons_absorbed

wb = openpyxl.Workbook()
ws = wb.active

ws.cell(row=1, column=1).value = "Absorbed Photon Yield"
ws.cell(row=2, column=1).value = total_photons_absorbed
ws.cell(row=1, column=2).value = "Absorption Percentage"
ws.cell(row=2, column=2).value = round(perc_absorbed*100, 4)
ws.cell(row=1, column=3).value = "Quantum Yield Percentage"
ws.cell(row=2, column=3).value = round(plqy*100, 4)

timestr = time.strftime("%Y%m%d-%H%M%S")

wb.save("PLQY"+timestr+".xlsx")
